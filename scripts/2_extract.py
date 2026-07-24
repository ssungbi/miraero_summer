import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = ("모집자명", "계약상태", "보험료")
EXCLUDED_NAMES = {"한만희", "박지수", "박성빈", "김지훈", "서태환", "권우원"}


def find_header_row(worksheet, search_limit=20):
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=search_limit, values_only=True),
        start=1,
    ):
        values = {str(value).strip() for value in row if value not in (None, "")}
        if set(REQUIRED_COLUMNS).issubset(values):
            return row_number, [str(value).strip() if value is not None else "" for value in row]
    raise ValueError("필수 컬럼(모집자명, 계약상태, 보험료)이 포함된 헤더 행을 찾을 수 없습니다.")


def premium_as_int(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = str(value).replace(",", "").strip()
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


def process_data(file_path):
    source = Path(file_path).expanduser().resolve(strict=True)
    if source.suffix.lower() != ".xlsx":
        raise ValueError("입력 파일은 .xlsx 형식이어야 합니다.")

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row, headers = find_header_row(worksheet)
        positions = {name: headers.index(name) for name in REQUIRED_COLUMNS}
        totals = defaultdict(int)

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[positions["모집자명"]] or "").strip()
            status = str(row[positions["계약상태"]] or "").strip()
            if not name or name in EXCLUDED_NAMES or status != "정상":
                continue
            totals[name] += premium_as_int(row[positions["보험료"]])
    finally:
        workbook.close()

    ranked = sorted(totals.items(), key=lambda item: (-item[1], item[0]))[:20]
    return [
        {
            "rank": rank,
            "agent": agent,
            "premium_str": f"{premium:,}",
            "premium_raw": premium,
        }
        for rank, (agent, premium) in enumerate(ranked, start=1)
    ]


def write_data(data, output_path):
    destination = Path(output_path).expanduser().resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return destination


def parse_args():
    parser = argparse.ArgumentParser(
        description="검증된 OPENGA xlsx에서 미래로 써머 랭킹 데이터를 추출합니다."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="이번 실행에서 검증한 신계약 전체 xlsx 절대경로",
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).with_name("data.json")),
        help="생성할 data.json 경로",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    data = process_data(args.input)
    destination = write_data(data, args.output)
    print(
        json.dumps(
            {
                "status": "DATA_EXTRACTED",
                "source": str(Path(args.input).expanduser().resolve()),
                "output": str(destination),
                "rankedAgents": len(data),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
